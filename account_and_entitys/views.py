import numpy as np
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination

from budget_management.models import (
    get_entities_with_children,
    get_level_zero_children,
    get_zero_level_accounts,
    get_zero_level_projects,
)
from .models import (
    XX_Account,
    XX_Entity,
    XX_Project,
    XX_PivotFund,
    XX_TransactionAudit,
    XX_ACCOUNT_ENTITY_LIMIT,
    XX_BalanceReport,
    Account_Mapping,
    Budget_data,
    XX_ACCOUNT_mapping,
    XX_Entity_mapping,
    EnvelopeManager,
)
from .serializers import (
    AccountSerializer,
    EntitySerializer,
    PivotFundSerializer,
    ProjectSerializer,
    TransactionAuditSerializer,
    AccountEntityLimitSerializer,
    BalanceReportSerializer,
)
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status
import pandas as pd
from django.db import transaction
from .models import XX_ACCOUNT_ENTITY_LIMIT
from .serializers import AccountEntityLimitSerializer
from django.db.models import CharField
from django.db.models.functions import Cast
from django.db.models import Q
from .utils import get_oracle_report_data,get_mapping_for_fusion_data


class EntityPagination(PageNumberPagination):
    """Pagination class for entities and accounts"""

    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


# Account views
class AccountListView(APIView):
    """List all accounts with optional search"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        search_query = request.query_params.get("search", None)

        accounts = XX_Account.objects.all().order_by("account")

        accounts = get_zero_level_accounts(accounts)

        if search_query:
            # Cast account (int) to string for filtering
            accounts = accounts.filter(
                Q(
                    account__icontains=search_query
                )  # works because Django auto casts to text in SQL
            )

        serializer = AccountSerializer(accounts, many=True)

        return Response(
            {"message": "Accounts retrieved successfully.", "data": serializer.data}
        )


class AccountCreateView(APIView):
    """Create a new account"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = AccountSerializer(data=request.data)
        if serializer.is_valid():
            account = serializer.save()
            return Response(
                {
                    "message": "Account created successfully.",
                    "data": AccountSerializer(account).data,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"message": "Failed to create account.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class AccountDetailView(APIView):
    """Retrieve a specific account"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Account.objects.get(pk=pk)
        except XX_Account.DoesNotExist:
            return None

    def get(self, request, pk):
        account = self.get_object(pk)
        if account is None:
            return Response(
                {"message": "Account not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = AccountSerializer(account)
        return Response(
            {
                "message": "Account details retrieved successfully.",
                "data": serializer.data,
            }
        )


class AccountUpdateView(APIView):
    """Update a specific account"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Account.objects.get(pk=pk)
        except XX_Account.DoesNotExist:
            return None

    def put(self, request, pk):
        account = self.get_object(pk)
        if account is None:
            return Response(
                {"message": "Account not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = AccountSerializer(account, data=request.data)
        if serializer.is_valid():
            updated_account = serializer.save()
            return Response(
                {
                    "message": "Account updated successfully.",
                    "data": AccountSerializer(updated_account).data,
                }
            )
        return Response(
            {"message": "Failed to update account.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class AccountDeleteView(APIView):
    """Delete a specific account"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Account.objects.get(pk=pk)
        except XX_Account.DoesNotExist:
            return None

    def delete(self, request, pk):
        account = self.get_object(pk)
        if account is None:
            return Response(
                {"message": "Account not found."}, status=status.HTTP_404_NOT_FOUND
            )
        account.delete()
        return Response(
            {"message": "Account deleted successfully."}, status=status.HTTP_200_OK
        )


# Project views


class Upload_ProjectsView(APIView):
    """Upload projects via Excel file"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Handle file upload and process projects

        Expects an Excel file where the first sheet has rows like:
        ProjectCode | ParentCode | AliasDefault

        The view will upsert each row into `XX_Project`.
        """
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            # Use openpyxl to read Excel content (openpyxl is listed in requirements)
            from openpyxl import load_workbook
            from django.db import transaction

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            with transaction.atomic():
                # Iterate rows — assume header may be present; detect header by non-numeric first row
                first = True
                for row in sheet.iter_rows(values_only=True):
                    # Skip entirely empty rows
                    if not row or all(
                        [
                            c is None or (isinstance(c, str) and c.strip() == "")
                            for c in row
                        ]
                    ):
                        continue

                    # Normalize columns: project_code, parent_code, alias_default
                    project_code = str(row[0]).strip() if row[0] is not None else None
                    parent_code = (
                        str(row[1]).strip()
                        if len(row) > 1 and row[1] is not None
                        else None
                    )
                    alias_default = (
                        str(row[2]).strip()
                        if len(row) > 2 and row[2] is not None
                        else None
                    )

                    # If the first row looks like a header (non-numeric project code) we skip it
                    if first:
                        first = False
                        header_like = False
                        # treat as header if first cell contains non-digit characters and not a typical code
                        if project_code and not any(
                            ch.isdigit() for ch in project_code
                        ):
                            header_like = True
                        if header_like:
                            continue

                    # Validate project_code
                    if not project_code:
                        skipped += 1
                        continue

                    # Upsert: update if exists, else create
                    try:
                        obj, created_flag = XX_Project.objects.update_or_create(
                            project=project_code,
                            defaults={
                                "parent": parent_code,
                                "alias_default": alias_default,
                            },
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {"project_code": project_code, "error": str(row_err)}
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }

            return Response(
                {"status": "ok", "summary": summary}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class Upload_AccountsView(APIView):
    """Upload accounts via Excel file"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Handle file upload and process accounts

        Expects an Excel file where the first sheet has rows like:
        AccountCode | ParentCode | AliasDefault

        The view will upsert each row into `XX_Account`.
        """
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            with transaction.atomic():
                first = True
                for row in sheet.iter_rows(values_only=True):
                    if not row or all(
                        [
                            c is None or (isinstance(c, str) and c.strip() == "")
                            for c in row
                        ]
                    ):
                        continue

                    account_code = str(row[0]).strip() if row[0] is not None else None
                    parent_code = (
                        str(row[1]).strip()
                        if len(row) > 1 and row[1] is not None
                        else None
                    )
                    alias_default = (
                        str(row[2]).strip()
                        if len(row) > 2 and row[2] is not None
                        else None
                    )

                    if first:
                        first = False
                        header_like = False
                        if account_code and not any(
                            ch.isdigit() for ch in account_code
                        ):
                            header_like = True
                        if header_like:
                            continue

                    if not account_code:
                        skipped += 1
                        continue

                    try:
                        obj, created_flag = XX_Account.objects.update_or_create(
                            account=account_code,
                            defaults={
                                "parent": parent_code,
                                "alias_default": alias_default,
                            },
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {"account_code": account_code, "error": str(row_err)}
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }

            return Response(
                {"status": "ok", "summary": summary}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class Upload_EntitiesView(APIView):
    """Upload entities via Excel file"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Handle file upload and process entities

        Expects an Excel file where the first sheet has rows like:
        EntityCode | ParentCode | AliasDefault

        The view will upsert each row into `XX_Entity`.
        """
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            with transaction.atomic():
                first = True
                for row in sheet.iter_rows(values_only=True):
                    if not row or all(
                        [
                            c is None or (isinstance(c, str) and c.strip() == "")
                            for c in row
                        ]
                    ):
                        continue

                    entity_code = str(row[0]).strip() if row[0] is not None else None
                    parent_code = (
                        str(row[1]).strip()
                        if len(row) > 1 and row[1] is not None
                        else None
                    )
                    alias_default = (
                        str(row[2]).strip()
                        if len(row) > 2 and row[2] is not None
                        else None
                    )

                    if first:
                        first = False
                        header_like = False
                        if entity_code and not any(ch.isdigit() for ch in entity_code):
                            header_like = True
                        if header_like:
                            continue

                    if not entity_code:
                        skipped += 1
                        continue

                    try:
                        obj, created_flag = XX_Entity.objects.update_or_create(
                            entity=entity_code,
                            defaults={
                                "parent": parent_code,
                                "alias_default": alias_default,
                            },
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {"entity_code": entity_code, "error": str(row_err)}
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }

            return Response(
                {"status": "ok", "summary": summary}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class UploadAccountMappingView(APIView):
    """Upload account mapping entries via Excel file"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        """Handle file upload and persist account mappings"""
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            import math

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            duplicates = 0
            skipped = 0
            errors = []

            def normalize(value):
                if value is None:
                    return None
                if isinstance(value, float):
                    if math.isnan(value):
                        return None
                    if value.is_integer():
                        return str(int(value))
                return str(value).strip()

            with transaction.atomic():
                first_row = True
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if not row or all(
                        cell is None or (isinstance(cell, str) and cell.strip() == "")
                        for cell in row
                    ):
                        continue

                    source_account = normalize(row[0]) if len(row) > 0 else None
                    target_account = normalize(row[1]) if len(row) > 1 else None

                    if first_row:
                        first_row = False
                        lower_source = (source_account or "").lower()
                        lower_target = (target_account or "").lower()
                        if lower_source in {
                            "source",
                            "source_account",
                        } or lower_target in {"target", "target_account"}:
                            continue

                    if not source_account or not target_account:
                        skipped += 1
                        continue

                    try:
                        _, created_flag = Account_Mapping.objects.get_or_create(
                            source_account=source_account,
                            target_account=target_account,
                        )
                        if created_flag:
                            created += 1
                        else:
                            duplicates += 1
                    except Exception as row_err:
                        errors.append(
                            {
                                "row": idx,
                                "source_account": source_account,
                                "target_account": target_account,
                                "error": str(row_err),
                            }
                        )

            summary = {
                "created": created,
                "duplicates": duplicates,
                "skipped": skipped,
                "errors": errors,
            }

            return Response(
                {"status": "ok", "summary": summary},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class ProjectListView(APIView):
    """List all Projects with optional search"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        search_query = request.query_params.get("search", None)

        if not request.user.projects.exists():
            projects = XX_Project.objects.all().order_by("project")
        else:
            projects = request.user.projects.all().order_by("project")
            all_projects = []
            all_projects.extend([proj.project for proj in projects])
            for proj in projects:
                all_projects.extend(
                    EnvelopeManager.get_all_children(
                        XX_Project.objects.all(), proj.project
                    )
                )
            projects = XX_Project.objects.filter(project__in=all_projects)
        # projects = get_zero_level_projects(projects)

        if search_query:
            # Cast project (int) to string for filtering
            projects = projects.filter(
                Q(
                    project__icontains=search_query
                )  # works because Django auto casts to text in SQL
            )

        serializer = ProjectSerializer(projects, many=True)

        return Response(
            {"message": "Projects retrieved successfully.", "data": serializer.data}
        )


class ProjectEnvelopeListView(APIView):
    """List all Projects with optional search"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        search_query = request.query_params.get("search", None)

        if not request.user.projects.exists():
            projects = XX_Project.objects.all().order_by("project")
            envelope_projects = []
            for proj in projects:
                if EnvelopeManager.Has_Envelope(proj.project):
                    envelope_projects.append(proj.project)
            projects = XX_Project.objects.filter(project__in=envelope_projects)
        else:
            projects = request.user.projects.all().order_by("project")
            all_projects = []
            all_projects.extend([proj.project for proj in projects])
            for proj in projects:
                all_projects.extend(
                    EnvelopeManager.get_all_children(
                        XX_Project.objects.all(), proj.project
                    )
                )
            envelope_projects = []
            for proj in all_projects:
                if EnvelopeManager.Has_Envelope(proj):
                    envelope_projects.append(proj)
            projects = XX_Project.objects.filter(project__in=envelope_projects)
        # projects = get_zero_level_projects(projects)

        if search_query:
            # Cast project (int) to string for filtering
            projects = projects.filter(
                Q(
                    project__icontains=search_query
                )  # works because Django auto casts to text in SQL
            )

        serializer = ProjectSerializer(projects, many=True)

        return Response(
            {"message": "Projects retrieved successfully.", "data": serializer.data}
        )


class ProjectCreateView(APIView):
    """Create a new project"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = ProjectSerializer(data=request.data)
        if serializer.is_valid():
            project = serializer.save()
            return Response(
                {
                    "message": "Project created successfully.",
                    "data": ProjectSerializer(project).data,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"message": "Failed to create project.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class ProjectDetailView(APIView):
    """Retrieve a specific project"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Project.objects.get(pk=pk)
        except XX_Project.DoesNotExist:
            return None

    def get(self, request, pk):
        project = self.get_object(pk)
        if project is None:
            return Response(
                {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = ProjectSerializer(project)
        return Response(
            {
                "message": "Project details retrieved successfully.",
                "data": serializer.data,
            }
        )


class ProjectUpdateView(APIView):
    """Update a specific project"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Project.objects.get(pk=pk)
        except XX_Project.DoesNotExist:
            return None

    def put(self, request, pk):
        project = self.get_object(pk)
        if project is None:
            return Response(
                {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = ProjectSerializer(project, data=request.data)
        if serializer.is_valid():
            updated_project = serializer.save()
            return Response(
                {
                    "message": "Project updated successfully.",
                    "data": ProjectSerializer(updated_project).data,
                }
            )
        return Response(
            {"message": "Failed to update project.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class ProjectDeleteView(APIView):
    """Delete a specific project"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Project.objects.get(pk=pk)
        except XX_Project.DoesNotExist:
            return None

    def delete(self, request, pk):
        project = self.get_object(pk)
        if project is None:
            return Response(
                {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
            )
        project.delete()
        return Response(
            {"message": "Project deleted successfully."}, status=status.HTTP_200_OK
        )


# Entity views
class EntityListView(APIView):
    """List all entities"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        # 🔹 Apply permissions filter
        if request.user.abilities.count() > 0:
            entity_ids = [
                ability.Entity.id
                for ability in request.user.abilities.all()
                if ability.Entity
            ]
            # Get all accessible entities including their children
            entities = get_entities_with_children(entity_ids)
        else:
            # If no permissions filter, get all entities
            entities = list(XX_Entity.objects.all().order_by("entity"))

        # 🔹 Get only level zero children from the accessible entities
        level_zero_entities = get_level_zero_children([e.id for e in entities])

        # 🔹 Apply search filter
        search_query = request.query_params.get("search")
        if search_query:
            search_lower = search_query.lower()
            level_zero_entities = [
                e for e in level_zero_entities if search_lower in str(e.entity).lower()
            ]

        # 🔹 Sort the final results
        level_zero_entities.sort(key=lambda x: x.entity)

        serializer = EntitySerializer(level_zero_entities, many=True)
        return Response(
            {"message": "Accounts retrieved successfully.", "data": serializer.data}
        )


class EntityMappingListView(APIView):
    """List all entity mappings"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Implement logic to retrieve and return entity mappings
        data = get_mapping_for_fusion_data()
        return Response({"message": "Entity mappings retrieved successfully.", "data": data})


class EntityCreateView(APIView):
    """Create a new entity"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = EntitySerializer(data=request.data)
        if serializer.is_valid():
            entity = serializer.save()
            return Response(
                {
                    "message": "Entity created successfully.",
                    "data": EntitySerializer(entity).data,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"message": "Failed to create entity.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class EntityDetailView(APIView):
    """Retrieve a specific entity"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Entity.objects.get(pk=pk)
        except XX_Entity.DoesNotExist:
            return None

    def get(self, request, pk):
        entity = self.get_object(pk)
        if entity is None:
            return Response(
                {"message": "Entity not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = EntitySerializer(entity)
        return Response(
            {
                "message": "Entity details retrieved successfully.",
                "data": serializer.data,
            }
        )


class EntityUpdateView(APIView):
    """Update a specific entity"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Entity.objects.get(pk=pk)
        except XX_Entity.DoesNotExist:
            return None

    def put(self, request, pk):
        entity = self.get_object(pk)
        if entity is None:
            return Response(
                {"message": "Entity not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = EntitySerializer(entity, data=request.data)
        if serializer.is_valid():
            updated_entity = serializer.save()
            return Response(
                {
                    "message": "Entity updated successfully.",
                    "data": EntitySerializer(updated_entity).data,
                }
            )
        return Response(
            {"message": "Failed to update entity.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class EntityDeleteView(APIView):
    """Delete a specific entity"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_Entity.objects.get(pk=pk)
        except XX_Entity.DoesNotExist:
            return None

    def delete(self, request, pk):
        entity = self.get_object(pk)
        if entity is None:
            return Response(
                {"message": "Entity not found."}, status=status.HTTP_404_NOT_FOUND
            )
        entity.delete()
        return Response(
            {"message": "Entity deleted successfully."}, status=status.HTTP_200_OK
        )


# PivotFund views
class PivotFundListView(APIView):
    """List all pivot funds"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        # Allow filtering by entity, account, and year
        entity_id = request.query_params.get("entity")
        account_id = request.query_params.get("account")
        project_id = request.query_params.get("project")
        year = request.query_params.get("year")

        pivot_funds = XX_PivotFund.objects.all()

        if entity_id:
            pivot_funds = pivot_funds.filter(entity=entity_id)
        if account_id:
            pivot_funds = pivot_funds.filter(account=account_id)
        if project_id:
            pivot_funds = pivot_funds.filter(project=project_id)
        if year:
            pivot_funds = pivot_funds.filter(year=year)

        # Order by year, entity, account
        pivot_funds = pivot_funds.order_by(
            "-year", "entity__entity", "account__account", "project__project"
        )

        # Handle pagination
        paginator = self.pagination_class()
        paginated_funds = paginator.paginate_queryset(pivot_funds, request)
        serializer = PivotFundSerializer(paginated_funds, many=True)

        return paginator.get_paginated_response(serializer.data)


class PivotFundCreateView(APIView):
    """Create a new pivot fund"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        # Handle batch creation
        if isinstance(request.data, list):
            created_funds = []
            errors = []

            for index, fund_data in enumerate(request.data):
                serializer = PivotFundSerializer(data=fund_data)
                if serializer.is_valid():
                    fund = serializer.save()
                    created_funds.append(PivotFundSerializer(fund).data)
                else:
                    errors.append(
                        {"index": index, "errors": serializer.errors, "data": fund_data}
                    )

            response_data = {
                "message": f"Created {len(created_funds)} pivot funds, with {len(errors)} errors.",
                "created": created_funds,
                "errors": errors,
            }

            if errors and not created_funds:
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)
            elif errors:
                return Response(response_data, status=status.HTTP_207_MULTI_STATUS)
            else:
                return Response(response_data, status=status.HTTP_201_CREATED)

        # Handle single creation
        else:
            serializer = PivotFundSerializer(data=request.data)
            if serializer.is_valid():
                fund = serializer.save()
                return Response(
                    {
                        "message": "Pivot fund created successfully.",
                        "data": PivotFundSerializer(fund).data,
                    },
                    status=status.HTTP_201_CREATED,
                )
            return Response(
                {
                    "message": "Failed to create pivot fund.",
                    "errors": serializer.errors,
                },
                status=status.HTTP_400_BAD_REQUEST,
            )


class PivotFundDetailView(APIView):
    """Retrieve a specific pivot fund"""

    permission_classes = [IsAuthenticated]

    def get_object(self, entity, account, project):
        try:

            return XX_PivotFund.objects.get(
                entity=entity, account=account, project=project
            )

        except XX_PivotFund.DoesNotExist:

            return None

    def get(self, request):

        entity = request.query_params.get("entity_id")
        account = request.query_params.get("account_id")
        project = request.query_params.get("project_id")
        print(entity, account, project)
        pivot_fund = self.get_object(entity, account, project)

        if pivot_fund is None:
            return Response(
                {"message": "Pivot fund not found."}, status=status.HTTP_200_OK
            )
        serializer = PivotFundSerializer(pivot_fund)
        return Response(
            {
                "message": "Pivot fund details retrieved successfully.",
                "data": serializer.data,
            }
        )


class PivotFundUpdateView(APIView):
    """Update a specific pivot fund"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_PivotFund.objects.get(pk=pk)
        except XX_PivotFund.DoesNotExist:
            return None

    def put(self, request, pk):
        pivot_fund = self.get_object(pk)
        if pivot_fund is None:
            return Response(
                {"message": "Pivot fund not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = PivotFundSerializer(pivot_fund, data=request.data)
        if serializer.is_valid():
            updated_fund = serializer.save()
            return Response(
                {
                    "message": "Pivot fund updated successfully.",
                    "data": PivotFundSerializer(updated_fund).data,
                }
            )
        return Response(
            {"message": "Failed to update pivot fund.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class PivotFundDeleteView(APIView):
    """Delete a specific pivot fund"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_PivotFund.objects.get(pk=pk)
        except XX_PivotFund.DoesNotExist:
            return None

    def delete(self, request, pk):
        pivot_fund = self.get_object(pk)
        if pivot_fund is None:
            return Response(
                {"message": "Pivot fund not found."}, status=status.HTTP_404_NOT_FOUND
            )
        pivot_fund.delete()
        return Response(
            {"message": "Pivot fund deleted successfully."}, status=status.HTTP_200_OK
        )


# ADJD Transaction Audit views


class TransactionAuditListView(APIView):
    """List all ADJD transaction audit records"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        audit_records = XX_TransactionAudit.objects.all().order_by("-id")

        # Handle pagination
        paginator = self.pagination_class()
        paginated_records = paginator.paginate_queryset(audit_records, request)
        serializer = TransactionAuditSerializer(paginated_records, many=True)

        return paginator.get_paginated_response(serializer.data)


class TransactionAuditCreateView(APIView):
    """Create a new ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = TransactionAuditSerializer(data=request.data)
        if serializer.is_valid():
            audit_record = serializer.save()
            return Response(
                {
                    "message": "Audit record created successfully.",
                    "data": TransactionAuditSerializer(audit_record).data,
                },
                status=status.HTTP_201_CREATED,
            )
        return Response(
            {"message": "Failed to create audit record.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class TransactionAuditDetailView(APIView):
    """Retrieve a specific ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_TransactionAudit.objects.get(pk=pk)
        except XX_TransactionAudit.DoesNotExist:
            return None

    def get(self, request, pk):
        audit_record = self.get_object(pk)
        if audit_record is None:
            return Response(
                {"message": "Audit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = TransactionAuditSerializer(audit_record)
        return Response(
            {
                "message": "Audit record details retrieved successfully.",
                "data": serializer.data,
            }
        )


class TransactionAuditUpdateView(APIView):
    """Update a specific ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_TransactionAudit.objects.get(pk=pk)
        except XX_TransactionAudit.DoesNotExist:
            return None

    def put(self, request, pk):
        audit_record = self.get_object(pk)
        if audit_record is None:
            return Response(
                {"message": "Audit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = TransactionAuditSerializer(audit_record, data=request.data)
        if serializer.is_valid():
            updated_record = serializer.save()
            return Response(
                {
                    "message": "Audit record updated successfully.",
                    "data": TransactionAuditSerializer(updated_record).data,
                }
            )
        return Response(
            {"message": "Failed to update audit record.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class TransactionAuditDeleteView(APIView):
    """Delete a specific ADJD transaction audit record"""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_TransactionAudit.objects.get(pk=pk)
        except XX_TransactionAudit.DoesNotExist:
            return None

    def delete(self, request, pk):
        audit_record = self.get_object(pk)
        if audit_record is None:
            return Response(
                {"message": "Audit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        audit_record.delete()
        return Response(
            {"message": "Audit record deleted successfully."}, status=status.HTTP_200_OK
        )


class list_ACCOUNT_ENTITY_LIMIT(APIView):
    """List all ADJD transaction audit records"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        # Change "enity_id" to "entity_id"
        entity_id = request.query_params.get("cost_center")
        account_id = request.query_params.get("account_id")
        project_id = request.query_params.get("project_id")

        audit_records = XX_ACCOUNT_ENTITY_LIMIT.objects.filter(
            entity_id=entity_id
        ).order_by("-id")
        audit_records = audit_records.annotate(
            account_id_str=Cast("account_id", CharField())
        )
        audit_records = audit_records.annotate(
            project_id_str=Cast("project_id", CharField())
        )

        if account_id:
            audit_records = audit_records.filter(
                account_id_str__icontains=str(account_id)
            )
        if project_id:
            audit_records = audit_records.filter(
                project_id_str__icontains=str(project_id)
            )

        # Handle pagination
        paginator = self.pagination_class()
        paginated_records = paginator.paginate_queryset(audit_records, request)
        serializer = AccountEntityLimitSerializer(paginated_records, many=True)

        data = [
            {
                "id": record["id"],
                "account": record["account_id"],
                "project": record["project_id"],
                "is_transer_allowed_for_source": record[
                    "is_transer_allowed_for_source"
                ],
                "is_transer_allowed_for_target": record[
                    "is_transer_allowed_for_target"
                ],
                "is_transer_allowed": record["is_transer_allowed"],
                "source_count": record["source_count"],
                "target_count": record["target_count"],
            }
            for record in serializer.data
        ]

        return paginator.get_paginated_response(data)


class AccountEntityLimitAPI(APIView):
    """Handle both listing and creation of ACCOUNT_ENTITY_LIMIT records"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]  # For file upload support

    def get(self, request):
        """List all records with optional filtering by cost_center"""
        entity_id = request.query_params.get("cost_center")

        audit_records = XX_ACCOUNT_ENTITY_LIMIT.objects.filter(
            entity_id=entity_id
        ).order_by("-id")

        paginator = self.pagination_class()
        paginated_records = paginator.paginate_queryset(audit_records, request)
        serializer = AccountEntityLimitSerializer(paginated_records, many=True)

        data = [
            {
                "id": record["id"],
                "account": record["account_id"],
                "project": record["project_id"],
                "is_transfer_allowed_for_source": record[
                    "is_transfer_allowed_for_source"
                ],
                "is_transfer_allowed_for_target": record[
                    "is_transfer_allowed_for_target"
                ],
                "is_transfer_allowed": record["is_transfer_allowed"],
                "source_count": record["source_count"],
                "target_count": record["target_count"],
            }
            for record in serializer.data
        ]

        return paginator.get_paginated_response(data)

    def post(self, request):
        """Handle both single record creation and bulk upload via file"""
        # Check if file is present for bulk upload
        uploaded_file = request.FILES.get("file")

        if uploaded_file:
            return self._handle_file_upload(uploaded_file)
        else:
            return self._handle_single_record(request.data)

    def _handle_file_upload(self, file):
        """Process Excel file for bulk creation"""
        try:
            # Read Excel file
            df = pd.read_excel(file)

            # Clean column names (convert to lowercase and strip whitespace)
            df.columns = df.columns.str.strip().str.lower()
            df = df.replace([np.nan, pd.NA, pd.NaT, "", "NULL", "null"], None)

            # Convert to list of dictionaries
            records = df.to_dict("records")

            created_count = 0
            errors = []

            with transaction.atomic():
                for idx, record in enumerate(records, start=1):
                    try:
                        serializer = AccountEntityLimitSerializer(data=record)
                        if serializer.is_valid():
                            serializer.save()
                            created_count += 1
                        else:
                            errors.append(
                                {
                                    "row": idx,
                                    "errors": serializer.errors,
                                    "data": record,
                                }
                            )
                    except Exception as e:
                        errors.append({"row": idx, "error": str(e), "data": record})

            response = {
                "status": "success",
                "created_count": created_count,
                "error_count": len(errors),
                "errors": errors if errors else None,
            }

            return Response(
                response,
                status=(
                    status.HTTP_201_CREATED
                    if created_count
                    else status.HTTP_400_BAD_REQUEST
                ),
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )

    def _handle_single_record(self, data):
        """Handle single record creation"""
        serializer = AccountEntityLimitSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class UpdateAccountEntityLimit(APIView):
    """Update a specific account entity limit."""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_ACCOUNT_ENTITY_LIMIT.objects.get(pk=pk)
        except XX_ACCOUNT_ENTITY_LIMIT.DoesNotExist:
            return None

    def put(self, request):

        pk = request.query_params.get("pk")
        limit_record = self.get_object(pk)
        if limit_record is None:
            return Response(
                {"message": "Limit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        serializer = AccountEntityLimitSerializer(limit_record, data=request.data)
        if serializer.is_valid():
            updated_record = serializer.save()
            return Response(
                {
                    "message": "Limit record updated successfully.",
                    "data": AccountEntityLimitSerializer(updated_record).data,
                }
            )
        return Response(
            {"message": "Failed to update limit record.", "errors": serializer.errors},
            status=status.HTTP_400_BAD_REQUEST,
        )


class DeleteAccountEntityLimit(APIView):
    """Delete a specific account entity limit."""

    permission_classes = [IsAuthenticated]

    def get_object(self, pk):
        try:
            return XX_ACCOUNT_ENTITY_LIMIT.objects.get(pk=pk)
        except XX_ACCOUNT_ENTITY_LIMIT.DoesNotExist:
            return None

    def delete(self, request, pk):
        limit_record = self.get_object(pk)
        if limit_record is None:
            return Response(
                {"message": "Limit record not found."}, status=status.HTTP_404_NOT_FOUND
            )
        limit_record.delete()
        return Response(
            {"message": "Limit record deleted successfully."}, status=status.HTTP_200_OK
        )


# MainCurrency views


class RefreshBalanceReportView(APIView):
    """API view to refresh balance report data from Oracle"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Trigger balance report refresh"""
        from .utils import refresh_balance_report_data

        budget_name = request.data.get("control_budget_name", "MIC_HQ_MONTHLY")
        period_name = request.data.get("Period_name", "sep-25")

        try:
            print("Starting balance report refresh...")
            print(f"Budget: {budget_name}, Period: {period_name}")
            result = refresh_balance_report_data(budget_name, period_name)
            if result["success"]:
                return Response(
                    {
                        "success": True,
                        "message": result["message"],
                        "data": {
                            "created_count": result["details"].get("created_count", 0),
                            "deleted_count": result["details"].get("deleted_count", 0),
                            "error_count": result["details"].get("error_count", 0),
                            "budget_name": budget_name,
                        },
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                return Response(
                    {
                        "success": False,
                        "message": result["message"],
                        "errors": result.get("details", {}).get("errors", []),
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error refreshing balance report: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def get(self, request):
        """Get balance report refresh status"""
        from .models import XX_BalanceReport

        try:
            total_records = XX_BalanceReport.objects.count()
            latest_record = XX_BalanceReport.objects.order_by("-created_at").first()
            periods = list(
                XX_BalanceReport.objects.values_list(
                    "as_of_period", flat=True
                ).distinct()
            )

            return Response(
                {
                    "success": True,
                    "data": {
                        "total_records": total_records,
                        "available_periods": periods,
                        "latest_update": (
                            latest_record.created_at if latest_record else None
                        ),
                        "last_period": (
                            latest_record.as_of_period if latest_record else None
                        ),
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error getting balance report status: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportListView(APIView):
    """List balance report data with filtering"""

    permission_classes = [IsAuthenticated]
    pagination_class = EntityPagination

    def get(self, request):
        """Get balance report data with optional filtering"""
        from .models import XX_BalanceReport
        from .serializers import BalanceReportSerializer
        from .utils import extract_unique_segments_from_data

        try:
            control_budget_name = request.query_params.get("control_budget_name")
            period_name = request.query_params.get("as_of_period")

            # Check if user wants only unique segments
            extract_segments = (
                request.query_params.get("extract_segments", "").lower() == "true"
            )

            # Get data from Oracle
            data = get_oracle_report_data(control_budget_name, period_name)

            # Validate Oracle response structure
            if not isinstance(data, dict) or not data.get("success"):
                message = (
                    data.get("message")
                    if isinstance(data, dict)
                    else "Unexpected response"
                )
                return Response(
                    {
                        "success": False,
                        "message": f"Failed to retrieve report data: {message}",
                        "data": [],
                    },
                    status=status.HTTP_502_BAD_GATEWAY,
                )

            records = data.get("data", [])

            if extract_segments:
                unique_segments = extract_unique_segments_from_data(data)

                # Enrich segments with names (aliases); fallback to code when missing
                cost_centers = unique_segments.get("Cost_Center", []) or []
                accounts = unique_segments.get("Account", []) or []
                projects = unique_segments.get("Project", []) or []

                # Fetch aliases in bulk
                entity_alias_map = {
                    str(e.entity): (e.alias_default or str(e.entity))
                    for e in XX_Entity.objects.filter(entity__in=cost_centers)
                }
                account_alias_map = {
                    str(a.account): (a.alias_default or str(a.account))
                    for a in XX_Account.objects.filter(account__in=accounts)
                }
                project_alias_map = {
                    str(p.project): (p.alias_default or str(p.project))
                    for p in XX_Project.objects.filter(project__in=projects)
                }

                enriched = {
                    "Cost_Center": [
                        {
                            "code": str(code),
                            "name": entity_alias_map.get(str(code), str(code)),
                        }
                        for code in cost_centers
                    ],
                    "Account": [
                        {
                            "code": str(code),
                            "name": account_alias_map.get(str(code), str(code)),
                        }
                        for code in accounts
                    ],
                    "Project": [
                        {
                            "code": str(code),
                            "name": project_alias_map.get(str(code), str(code)),
                        }
                        for code in projects
                    ],
                    "total_records": unique_segments.get("total_records", 0),
                    "unique_combinations": unique_segments.get(
                        "unique_combinations", 0
                    ),
                }
                return Response(
                    {
                        "success": True,
                        "message": "Unique segments extracted successfully",
                        "data": enriched,
                    },
                    status=status.HTTP_200_OK,
                )

            # Otherwise, return the full data with unique segments included
            unique_segments = extract_unique_segments_from_data(data)

            # Enrich segments with names (aliases); fallback to code when missing
            cost_centers = unique_segments.get("Cost_Center", []) or []
            accounts = unique_segments.get("Account", []) or []
            projects = unique_segments.get("Project", []) or []

            entity_alias_map = {
                str(e.entity): (e.alias_default or str(e.entity))
                for e in XX_Entity.objects.filter(entity__in=cost_centers)
            }
            account_alias_map = {
                str(a.account): (a.alias_default or str(a.account))
                for a in XX_Account.objects.filter(account__in=accounts)
            }
            project_alias_map = {
                str(p.project): (p.alias_default or str(p.project))
                for p in XX_Project.objects.filter(project__in=projects)
            }

            enriched_unique_segments = {
                "Cost_Center": [
                    {
                        "code": str(code),
                        "name": entity_alias_map.get(str(code), str(code)),
                    }
                    for code in cost_centers
                ],
                "Account": [
                    {
                        "code": str(code),
                        "name": account_alias_map.get(str(code), str(code)),
                    }
                    for code in accounts
                ],
                "Project": [
                    {
                        "code": str(code),
                        "name": project_alias_map.get(str(code), str(code)),
                    }
                    for code in projects
                ],
                "total_records": unique_segments.get("total_records", 0),
                "unique_combinations": unique_segments.get("unique_combinations", 0),
            }

            return Response(
                {
                    "success": True,
                    "message": "Balance report data retrieved successfully",
                    "data": {
                        "records": records,
                        "unique_segments": enriched_unique_segments,
                        "total_records": len(records),
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error retrieving balance report data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportSegmentsView(APIView):
    """API to get all unique segments from balance report"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get unique values for segment1, segment2, and segment3"""
        from .models import XX_BalanceReport

        try:
            # Get unique segments with filters
            segment1_filter = request.query_params.get("segment1")
            segment2_filter = request.query_params.get("segment2")

            queryset = XX_BalanceReport.objects.all()

            # Apply filters if provided
            if segment1_filter:
                queryset = queryset.filter(segment1=segment1_filter)
            if segment2_filter:
                queryset = queryset.filter(segment2=segment2_filter)

            # Get unique values for each segment
            segment1_values = list(
                XX_BalanceReport.objects.filter(segment1__isnull=False)
                .values_list("segment1", flat=True)
                .distinct()
                .order_by("segment1")
            )

            segment2_values = list(
                queryset.filter(segment2__isnull=False)
                .values_list("segment2", flat=True)
                .distinct()
                .order_by("segment2")
            )

            segment3_values = list(
                queryset.filter(segment3__isnull=False)
                .values_list("segment3", flat=True)
                .distinct()
                .order_by("segment3")
            )

            return Response(
                {
                    "success": True,
                    "data": {
                        "segment1": segment1_values,
                        "segment2": segment2_values,
                        "segment3": segment3_values,
                        "total_combinations": queryset.count(),
                    },
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"success": False, "message": f"Error retrieving segments: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportOracleSegmentsView(APIView):
    """API to get unique segments from Oracle balance report data"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get unique segments from Oracle data with optional filtering"""
        from .utils import extract_unique_segments_from_data

        try:
            # Get filter parameters
            control_budget_name = request.query_params.get("control_budget_name")
            period_name = request.query_params.get("as_of_period")
            segment1_filter = request.query_params.get("segment1")
            segment2_filter = request.query_params.get("segment2")
            segment3_filter = request.query_params.get("segment3")

            # Get data from Oracle with filters
            oracle_data = get_oracle_report_data(
                control_budget_name,
                period_name,
                segment1_filter,
                segment2_filter,
                segment3_filter,
            )

            if not oracle_data:
                return Response(
                    {
                        "success": False,
                        "message": "No data found with the specified filters",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Extract unique segments from the Oracle data
            unique_segments = extract_unique_segments_from_data(oracle_data)

            return Response(
                {
                    "success": True,
                    "message": f"Successfully extracted unique segments from {len(oracle_data)} records",
                    "data": unique_segments,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error extracting segments from Oracle data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        """Extract unique segments from provided balance report data"""
        from .utils import extract_unique_segments_from_data

        try:
            # Get data from request body
            balance_data = request.data.get("data", [])

            if not balance_data:
                return Response(
                    {
                        "success": False,
                        "message": "No balance report data provided in request body",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Extract unique segments from the provided data
            unique_segments = extract_unique_segments_from_data(balance_data)

            return Response(
                {
                    "success": True,
                    "message": f"Successfully extracted unique segments from {len(balance_data)} records",
                    "data": unique_segments,
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error extracting segments from provided data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class BalanceReportFinancialDataView(APIView):
    """API to get financial data for specific segment combination"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        """Get financial data for a specific segment1, segment2, segment3 combination"""
        from .models import XX_BalanceReport
        from django.db.models import Sum, Avg, Count

        try:
            from .utils import refresh_balance_report_data

            budget_name = request.data.get("control_budget_name", "MIC_HQ_MONTHLY")
            period_name = request.data.get("Period_name", "sep-25")

            try:
                result = refresh_balance_report_data(
                    budget_name, period_name=period_name
                )
            except Exception as e:
                return Response(
                    {
                        "success": False,
                        "message": f"Error refreshing balance report data: {str(e)}",
                    },
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR,
                )

            # Get segments from query parameters
            segment1 = request.query_params.get("segment1")
            segment2 = request.query_params.get("segment2")
            segment3 = request.query_params.get("segment3")

            # Validate that all segments are provided
            if not all([segment1, segment2, segment3]):
                return Response(
                    {
                        "success": False,
                        "message": "All three segments (segment1, segment2, segment3) are required",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Filter records by the segments
            queryset = XX_BalanceReport.objects.filter(
                segment1=segment1, segment2=segment2, segment3=segment3
            )

            if not queryset.exists():
                return Response(
                    {
                        "success": False,
                        "message": f"No data found for segments: {segment1}/{segment2}/{segment3}",
                    },
                    status=status.HTTP_404_NOT_FOUND,
                )

            # Get the most recent record for this combination
            latest_record = queryset.order_by("-created_at").first()

            # Calculate aggregated data if multiple records exist
            aggregated_data = queryset.aggregate(
                total_actual_ytd=Sum("actual_ytd"),
                total_encumbrance_ytd=Sum("encumbrance_ytd"),
                total_funds_available=Sum("funds_available_asof"),
                total_other_ytd=Sum("other_ytd"),
                total_budget_ytd=Sum("budget_ytd"),
                avg_actual_ytd=Avg("actual_ytd"),
                avg_encumbrance_ytd=Avg("encumbrance_ytd"),
                avg_funds_available=Avg("funds_available_asof"),
                record_count=Count("id"),
            )

            # Prepare response data
            financial_data = {
                "segments": {
                    "segment1": segment1,
                    "segment2": segment2,
                    "segment3": segment3,
                },
                "latest_record": {
                    "control_budget_name": latest_record.control_budget_name,
                    "ledger_name": latest_record.ledger_name,
                    "as_of_period": latest_record.as_of_period,
                    "actual_ytd": (
                        float(latest_record.actual_ytd)
                        if latest_record.actual_ytd
                        else 0
                    ),
                    "encumbrance_ytd": (
                        float(latest_record.encumbrance_ytd)
                        if latest_record.encumbrance_ytd
                        else 0
                    ),
                    "funds_available_asof": (
                        float(latest_record.funds_available_asof)
                        if latest_record.funds_available_asof
                        else 0
                    ),
                    "other_ytd": (
                        float(latest_record.other_ytd) if latest_record.other_ytd else 0
                    ),
                    "budget_ytd": (
                        float(latest_record.budget_ytd)
                        if latest_record.budget_ytd
                        else 0
                    ),
                    "last_updated": latest_record.created_at,
                },
                "aggregated_totals": {
                    "total_actual_ytd": float(aggregated_data["total_actual_ytd"] or 0),
                    "total_encumbrance_ytd": float(
                        aggregated_data["total_encumbrance_ytd"] or 0
                    ),
                    "total_funds_available": float(
                        aggregated_data["total_funds_available"] or 0
                    ),
                    "total_other_ytd": float(aggregated_data["total_other_ytd"] or 0),
                    "total_budget_ytd": float(aggregated_data["total_budget_ytd"] or 0),
                    "record_count": aggregated_data["record_count"],
                },
                "calculated_metrics": {
                    "budget_utilization_percent": (
                        round(
                            (
                                float(aggregated_data["total_actual_ytd"] or 0)
                                / float(aggregated_data["total_budget_ytd"] or 1)
                            )
                            * 100,
                            2,
                        )
                        if aggregated_data["total_budget_ytd"]
                        else 0
                    ),
                    "funds_remaining": float(
                        aggregated_data["total_funds_available"] or 0
                    ),
                    "total_committed": float(aggregated_data["total_actual_ytd"] or 0)
                    + float(aggregated_data["total_encumbrance_ytd"] or 0),
                },
            }

            return Response(
                {"success": True, "data": financial_data}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error retrieving financial data: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

    def post(self, request):
        """Get financial data for multiple segment combinations"""
        from .models import XX_BalanceReport

        try:
            segment_combinations = request.data.get("segments", [])

            if not segment_combinations:
                return Response(
                    {
                        "success": False,
                        "message": 'Please provide segment combinations in the format: [{"segment1": "10001", "segment2": "2205403", "segment3": "CTRLCE1"}]',
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            results = []

            for combo in segment_combinations:
                segment1 = combo.get("segment1")
                segment2 = combo.get("segment2")
                segment3 = combo.get("segment3")

                if not all([segment1, segment2, segment3]):
                    results.append(
                        {
                            "segments": combo,
                            "success": False,
                            "message": "Missing segment values",
                        }
                    )
                    continue

                # Get data for this combination
                record = (
                    XX_BalanceReport.objects.filter(
                        segment1=segment1, segment2=segment2, segment3=segment3
                    )
                    .order_by("-created_at")
                    .first()
                )

                if record:
                    results.append(
                        {
                            "segments": {
                                "segment1": segment1,
                                "segment2": segment2,
                                "segment3": segment3,
                            },
                            "success": True,
                            "data": {
                                "actual_ytd": (
                                    float(record.actual_ytd) if record.actual_ytd else 0
                                ),
                                "encumbrance_ytd": (
                                    float(record.encumbrance_ytd)
                                    if record.encumbrance_ytd
                                    else 0
                                ),
                                "funds_available_asof": (
                                    float(record.funds_available_asof)
                                    if record.funds_available_asof
                                    else 0
                                ),
                                "other_ytd": (
                                    float(record.other_ytd) if record.other_ytd else 0
                                ),
                                "budget_ytd": (
                                    float(record.budget_ytd) if record.budget_ytd else 0
                                ),
                                "as_of_period": record.as_of_period,
                                "last_updated": record.created_at,
                            },
                        }
                    )
                else:
                    results.append(
                        {
                            "segments": combo,
                            "success": False,
                            "message": "No data found for this segment combination",
                        }
                    )

            return Response(
                {
                    "success": True,
                    "data": results,
                    "total_requested": len(segment_combinations),
                    "found": len([r for r in results if r["success"]]),
                },
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {
                    "success": False,
                    "message": f"Error processing segment combinations: {str(e)}",
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class Single_BalanceReportView(APIView):
    """Retrieve a specific balance report record by ID"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        control_budget_name = request.query_params.get("control_budget_name")
        period_name = request.query_params.get("as_of_period")
        segment1 = request.query_params.get("segment1")
        segment2 = request.query_params.get("segment2")
        segment3 = request.query_params.get("segment3")

        data = get_oracle_report_data(
            control_budget_name, period_name, segment1, segment2, segment3
        )

        if data is None:
            return Response(
                {"message": "Balance report record not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        return Response(
            {
                "message": "Balance report record details retrieved successfully.",
                "data": data,
            }
        )


# envlop
class Upload_ProjectEnvelopeView(APIView):
    """Upload project envelopes via Excel file"""

    permission_classes = [IsAuthenticated]

    def post(self, request):
        """Handle file upload and process project envelopes

        Expects an Excel file where the first sheet has rows like:
        ProjectCodeWithAlias | EnvelopeNumber

        The view will upsert each row into `Project_Envelope`.
        """
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction
            from decimal import Decimal, InvalidOperation
            import math
            import re

            from .models import Project_Envelope

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            def extract_project_code(raw_value):
                """Return the leading numeric project code from the provided cell."""
                if raw_value is None:
                    return None
                if isinstance(raw_value, int):
                    return str(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return None
                    if raw_value.is_integer():
                        return str(int(raw_value))
                    return str(raw_value)
                value_str = str(raw_value).strip()
                match = re.match(r"^\s*(\d+)", value_str)
                return match.group(1) if match else None

            with transaction.atomic():
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if not row or all(
                        cell is None or (isinstance(cell, str) and cell.strip() == "")
                        for cell in row
                    ):
                        continue

                    project_code = extract_project_code(
                        row[0] if len(row) > 0 else None
                    )
                    envelope_raw = row[1] if len(row) > 1 else None

                    if not project_code:
                        skipped += 1
                        continue

                    if envelope_raw is None or (
                        isinstance(envelope_raw, str) and envelope_raw.strip() == ""
                    ):
                        skipped += 1
                        continue

                    try:
                        if isinstance(envelope_raw, str):
                            cleaned = re.sub(
                                r"[^0-9.\-]", "", envelope_raw.replace(",", "")
                            )
                            envelope_val = Decimal(cleaned) if cleaned != "" else None
                        else:
                            envelope_val = Decimal(envelope_raw)
                    except (InvalidOperation, TypeError):
                        errors.append(
                            {
                                "row": idx,
                                "project_code": project_code,
                                "error": f"Invalid envelope value: {envelope_raw}",
                            }
                        )
                        continue

                    if envelope_val is None:
                        skipped += 1
                        continue

                    try:
                        _, created_flag = Project_Envelope.objects.update_or_create(
                            project=project_code,
                            defaults={"envelope": envelope_val},
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {
                                "row": idx,
                                "project_code": project_code,
                                "error": str(row_err),
                            }
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }
            return Response(
                {"status": "ok", "summary": summary}, status=status.HTTP_200_OK
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class UploadBudgetDataView(APIView):
    """Upload budget data via Excel file"""

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def post(self, request):
        """Handle file upload and upsert budget data rows."""
        uploaded_file = request.FILES.get("file")

        if not uploaded_file:
            return Response(
                {"message": "No file uploaded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            from openpyxl import load_workbook
            from django.db import transaction
            from decimal import Decimal, InvalidOperation
            import math
            import re

            wb = load_workbook(filename=uploaded_file, read_only=True, data_only=True)
            sheet = wb.active

            created = 0
            updated = 0
            skipped = 0
            errors = []

            def extract_project_code(raw_value):
                """Return the project code from the provided cell."""
                if raw_value is None:
                    return None
                if isinstance(raw_value, int):
                    return str(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return None
                    if raw_value.is_integer():
                        return str(int(raw_value))
                    return str(raw_value)
                return str(raw_value).strip() or None

            def normalize_account(raw_value):
                if raw_value is None:
                    return None
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return None
                    if raw_value.is_integer():
                        return str(int(raw_value))
                    return str(raw_value)
                value_str = str(raw_value).strip()
                return value_str or None

            def parse_budget(raw_value, label):
                if raw_value is None:
                    return Decimal("0")
                if isinstance(raw_value, int):
                    return Decimal(raw_value)
                if isinstance(raw_value, float):
                    if math.isnan(raw_value):
                        return Decimal("0")
                    return Decimal(str(raw_value))
                value_str = str(raw_value).strip()
                if not value_str:
                    return Decimal("0")
                cleaned = re.sub(r"[^0-9.\-]", "", value_str)
                if cleaned in {"", "-", ".", "-.", ".-"}:
                    return Decimal("0")
                try:
                    return Decimal(cleaned)
                except InvalidOperation as exc:
                    raise ValueError(f"Invalid {label} value: {raw_value}") from exc

            with transaction.atomic():
                for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if not row or all(
                        cell is None or (isinstance(cell, str) and cell.strip() == "")
                        for cell in row
                    ):
                        continue

                    project_code = extract_project_code(
                        row[0] if len(row) > 0 else None
                    )
                    account_code = normalize_account(row[1] if len(row) > 1 else None)

                    if not project_code or not account_code:
                        skipped += 1
                        continue

                    try:
                        fy24_budget = parse_budget(
                            row[2] if len(row) > 2 else None,
                            "FY24 budget",
                        )
                        fy25_budget = parse_budget(
                            row[3] if len(row) > 3 else None,
                            "FY25 budget",
                        )
                    except ValueError as budget_err:
                        errors.append(
                            {
                                "row": idx,
                                "project": project_code,
                                "account": account_code,
                                "error": str(budget_err),
                            }
                        )
                        continue

                    try:
                        _, created_flag = Budget_data.objects.update_or_create(
                            project=project_code,
                            account=account_code,
                            defaults={
                                "FY24_budget": fy24_budget,
                                "FY25_budget": fy25_budget,
                            },
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except Exception as row_err:
                        errors.append(
                            {
                                "row": idx,
                                "project": project_code,
                                "account": account_code,
                                "error": str(row_err),
                            }
                        )

            summary = {
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
            }
            return Response(
                {"status": "ok", "summary": summary},
                status=status.HTTP_200_OK,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_400_BAD_REQUEST,
            )


class ActiveProjectsWithEnvelopeView(APIView):
    """Return active projects with their current envelope and totals."""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        # Query params: year, month, IsApproved
        project_code = request.query_params.get("project_code", None)
        if project_code is None:
            return Response(
                {"message": "project_code query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        project = XX_Project.objects.get(pk=project_code)
        if project is None:
            return Response(
                {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
            )
        year = request.query_params.get("year", None)
        month = request.query_params.get("month", None)

        results = EnvelopeManager.Get_Current_Envelope_For_Project(
            project_code=project.project, year=year, month=month
        )

        if results and "project_totals" in results:
            # Transform into array of objects
            transformed_data = []

            # Extract data from the original structure
            for proj, data in results["project_totals"].items():
                transformed_data.append(
                    {
                        "project_code": proj,
                        "submitted_total": (
                            data["submitted"]["total"] if data["submitted"] else 0
                        ),
                        "approved_total": (
                            data["approved"]["total"] if data["approved"] else 0
                        ),
                    }
                )

            return Response(
                {
                    "message": "Active projects with envelope.",
                    "initial_envelope": results["initial_envelope"],
                    "current_envelope": results["current_envelope"],
                    "estimated_envelope": results["estimated_envelope"],
                    "data": transformed_data,
                }
            )


class ProjectWiseDashboardView(APIView):
    """Project-wise dashboard data"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        entity_code = request.query_params.get("entity_code", None)
        if entity_code is None:
            return Response(
                {"message": "entity_code query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        results = EnvelopeManager.Get_Dashboard_Data_For_Entity(entity_code)
        return Response(
            {
                "message": "Project-wise dashboard data.",
                "data": results,
            }
        )


class AccountWiseDashboardView(APIView):
    """Account-wise dashboard data"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        project_code = request.query_params.get("project_code", None)
        if project_code is None:
            return Response(
                {"message": "project_code query parameter is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        project = XX_Project.objects.get(pk=project_code)
        if project is None:
            return Response(
                {"message": "Project not found."}, status=status.HTTP_404_NOT_FOUND
            )
        results = EnvelopeManager.Get_Dashboard_Data_For_Project(project.project)
        return Response(
            {
                "message": "Project-wise dashboard data.",
                "data": results,
            }
        )


# Mapping

class UploadMappingExcelView(APIView):
    """
    Upload Excel file with Account and Entity mapping data.
    Expected sheets:
    - 'Account' sheet with columns: source_account, target_account
    - 'Entity' sheet with columns: source_entity, target_entity
    """

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser]

    def get(self, request):

        data = get_mapping_for_fusion_data()
        return Response(
            {
                "data": data
            }
        )

    def post(self, request):
        try:
            # Check if file is provided
            if "file" not in request.FILES:
                return Response(
                    {"status": "error", "message": "No file provided"},
                    status=status.HTTP_400_BAD_REQUEST,
                )

            uploaded_file = request.FILES["file"]

            # Validate file extension
            if not uploaded_file.name.endswith((".xlsx", ".xls")):
                return Response(
                    {
                        "status": "error",
                        "message": "File must be Excel format (.xlsx or .xls)",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

            # Read Excel file
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names

            results = {
                "account_mappings": {
                    "created": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": [],
                },
                "entity_mappings": {
                    "created": 0,
                    "updated": 0,
                    "skipped": 0,
                    "errors": [],
                },
            }

            # Process Account sheet
            if "Account" in sheet_names:
                try:
                    account_df = pd.read_excel(uploaded_file, sheet_name="Account")

                    # Validate required columns
                    required_cols = ["Source", "Target"]
                    missing_cols = [
                        col for col in required_cols if col not in account_df.columns
                    ]

                    if missing_cols:
                        results["account_mappings"]["errors"].append(
                            f"Missing columns in Account sheet: {', '.join(missing_cols)}"
                        )
                    else:
                        # Process each row
                        with transaction.atomic():
                            for index, row in account_df.iterrows():
                                try:
                                    source_account = str(row["Source"]).strip()
                                    target_account = str(row["Target"]).strip()

                                    # Skip rows with empty values
                                    if (
                                        pd.isna(row["Source"])
                                        or pd.isna(row["Target"])
                                        or source_account == ""
                                        or target_account == ""
                                    ):
                                        results["account_mappings"]["skipped"] += 1
                                        continue

                                    # Create or update account mapping
                                    obj, created = (
                                        XX_ACCOUNT_mapping.objects.update_or_create(
                                            source_account=source_account,
                                            target_account=target_account,
                                            defaults={"is_active": True},
                                        )
                                    )

                                    if created:
                                        results["account_mappings"]["created"] += 1
                                    else:
                                        results["account_mappings"]["updated"] += 1

                                except Exception as row_err:
                                    results["account_mappings"]["errors"].append(
                                        f"Row {index + 2}: {str(row_err)}"
                                    )

                except Exception as sheet_err:
                    results["account_mappings"]["errors"].append(
                        f"Error processing Account sheet: {str(sheet_err)}"
                    )
            else:
                results["account_mappings"]["errors"].append(
                    "Account sheet not found in Excel file"
                )

            # Process Entity sheet
            if "Entity" in sheet_names:
                try:
                    entity_df = pd.read_excel(uploaded_file, sheet_name="Entity")

                    # Validate required columns
                    required_cols = ["Source", "Target"]
                    missing_cols = [
                        col for col in required_cols if col not in entity_df.columns
                    ]

                    if missing_cols:
                        results["entity_mappings"]["errors"].append(
                            f"Missing columns in Entity sheet: {', '.join(missing_cols)}"
                        )
                    else:
                        # Process each row
                        with transaction.atomic():
                            for index, row in entity_df.iterrows():
                                try:
                                    source_entity = str(row["Source"]).strip()
                                    target_entity = str(row["Target"]).strip()

                                    # Skip rows with empty values
                                    if (
                                        pd.isna(row["Source"])
                                        or pd.isna(row["Target"])
                                        or source_entity == ""
                                        or target_entity == ""
                                    ):
                                        results["entity_mappings"]["skipped"] += 1
                                        continue

                                    # Create or update entity mapping
                                    obj, created = (
                                        XX_Entity_mapping.objects.update_or_create(
                                            source_entity=source_entity,
                                            target_entity=target_entity,
                                            defaults={"is_active": True},
                                        )
                                    )

                                    if created:
                                        results["entity_mappings"]["created"] += 1
                                    else:
                                        results["entity_mappings"]["updated"] += 1

                                except Exception as row_err:
                                    results["entity_mappings"]["errors"].append(
                                        f"Row {index + 2}: {str(row_err)}"
                                    )

                except Exception as sheet_err:
                    results["entity_mappings"]["errors"].append(
                        f"Error processing Entity sheet: {str(sheet_err)}"
                    )
            else:
                results["entity_mappings"]["errors"].append(
                    "Entity sheet not found in Excel file"
                )

            # Determine overall status
            has_errors = (
                len(results["account_mappings"]["errors"]) > 0
                or len(results["entity_mappings"]["errors"]) > 0
            )

            response_status = (
                status.HTTP_207_MULTI_STATUS if has_errors else status.HTTP_200_OK
            )

            return Response(
                {
                    "status": "completed_with_errors" if has_errors else "success",
                    "message": (
                        "File processed successfully"
                        if not has_errors
                        else "File processed with some errors"
                    ),
                    "results": results,
                },
                status=response_status,
            )

        except Exception as e:
            return Response(
                {"status": "error", "message": f"Unexpected error: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


# class AccountMappingListView(APIView):
#     """List all account mappings"""

#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         try:
#             mappings = XX_ACCOUNT_mapping.objects.filter(is_active=True)
#             from .serializers import AccountMappingSerializer

#             serializer = AccountMappingSerializer(mappings, many=True)
#             return Response(
#                 {
#                     "status": "success",
#                     "data": serializer.data,
#                     "count": mappings.count(),
#                 },
#                 status=status.HTTP_200_OK,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )


# class EntityMappingListView(APIView):
#     """List all entity mappings"""

#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         try:
#             mappings = XX_Entity_mapping.objects.filter(is_active=True)
#             from .serializers import EntityMappingSerializer

#             serializer = EntityMappingSerializer(mappings, many=True)
#             return Response(
#                 {
#                     "status": "success",
#                     "data": serializer.data,
#                     "count": mappings.count(),
#                 },
#                 status=status.HTTP_200_OK,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )


# class AccountMappingDetailView(APIView):
#     """Get, update, or delete specific account mapping"""

#     permission_classes = [IsAuthenticated]

#     def get(self, request, pk):
#         try:
#             mapping = XX_ACCOUNT_mapping.objects.get(id=pk)
#             from .serializers import AccountMappingSerializer

#             serializer = AccountMappingSerializer(mapping)
#             return Response(
#                 {"status": "success", "data": serializer.data},
#                 status=status.HTTP_200_OK,
#             )
#         except XX_ACCOUNT_mapping.DoesNotExist:
#             return Response(
#                 {"status": "error", "message": "Account mapping not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )

#     def put(self, request, pk):
#         try:
#             mapping = XX_ACCOUNT_mapping.objects.get(id=pk)
#             from .serializers import AccountMappingSerializer

#             serializer = AccountMappingSerializer(
#                 mapping, data=request.data, partial=True
#             )
#             if serializer.is_valid():
#                 serializer.save()
#                 return Response(
#                     {"status": "success", "data": serializer.data},
#                     status=status.HTTP_200_OK,
#                 )
#             return Response(
#                 {"status": "error", "errors": serializer.errors},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )
#         except XX_ACCOUNT_mapping.DoesNotExist:
#             return Response(
#                 {"status": "error", "message": "Account mapping not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )

#     def delete(self, request, pk):
#         try:
#             mapping = XX_ACCOUNT_mapping.objects.get(id=pk)
#             mapping.is_active = False
#             mapping.save()
#             return Response(
#                 {"status": "success", "message": "Account mapping deactivated"},
#                 status=status.HTTP_200_OK,
#             )
#         except XX_ACCOUNT_mapping.DoesNotExist:
#             return Response(
#                 {"status": "error", "message": "Account mapping not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )


# class EntityMappingDetailView(APIView):
#     """Get, update, or delete specific entity mapping"""

#     permission_classes = [IsAuthenticated]

#     def get(self, request, pk):
#         try:
#             mapping = XX_Entity_mapping.objects.get(id=pk)
#             from .serializers import EntityMappingSerializer

#             serializer = EntityMappingSerializer(mapping)
#             return Response(
#                 {"status": "success", "data": serializer.data},
#                 status=status.HTTP_200_OK,
#             )
#         except XX_Entity_mapping.DoesNotExist:
#             return Response(
#                 {"status": "error", "message": "Entity mapping not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )

#     def put(self, request, pk):
#         try:
#             mapping = XX_Entity_mapping.objects.get(id=pk)
#             from .serializers import EntityMappingSerializer

#             serializer = EntityMappingSerializer(
#                 mapping, data=request.data, partial=True
#             )
#             if serializer.is_valid():
#                 serializer.save()
#                 return Response(
#                     {"status": "success", "data": serializer.data},
#                     status=status.HTTP_200_OK,
#                 )
#             return Response(
#                 {"status": "error", "errors": serializer.errors},
#                 status=status.HTTP_400_BAD_REQUEST,
#             )
#         except XX_Entity_mapping.DoesNotExist:
#             return Response(
#                 {"status": "error", "message": "Entity mapping not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )

#     def delete(self, request, pk):
#         try:
#             mapping = XX_Entity_mapping.objects.get(id=pk)
#             mapping.is_active = False
#             mapping.save()
#             return Response(
#                 {"status": "success", "message": "Entity mapping deactivated"},
#                 status=status.HTTP_200_OK,
#             )
#         except XX_Entity_mapping.DoesNotExist:
#             return Response(
#                 {"status": "error", "message": "Entity mapping not found"},
#                 status=status.HTTP_404_NOT_FOUND,
#             )
#         except Exception as e:
#             return Response(
#                 {"status": "error", "message": str(e)},
#                 status=status.HTTP_500_INTERNAL_SERVER_ERROR,
#             )
