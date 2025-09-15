import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
from pathlib import Path
import time
from typing import Dict, List, Any
from test_upload_fbdi.zip_fbdi import excel_to_csv_and_zip
from test_upload_fbdi.upload_budget_fbdi import upload_budget_from_zip, upload_budget_fbdi_to_oracle

def create_clean_budget_template(template_path: str, output_path: str = None) -> str:
    """
    Create a clean copy of the BudgetImportTemplate with empty XCC_BUDGET_INTERFACE data rows.
    
    Args:
        template_path: Path to the original BudgetImportTemplate.xlsm
        output_path: Path for the clean template (optional, auto-generates if not provided)
    
    Returns:
        Path to the clean template file
    """
    template_file = Path(template_path)
    
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"BudgetImportTemplate_Clean_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)
    
    # Create a copy of the template
    shutil.copy2(template_file, output_path)
    print(f"Created clean template copy: {output_path}")
    
    # Open the copy and clear XCC_BUDGET_INTERFACE data
    wb = load_workbook(output_path)
    
    if "XCC_BUDGET_INTERFACE" in wb.sheetnames:
        budget_sheet = wb["XCC_BUDGET_INTERFACE"]
        
        # Keep rows 1-4 (instructions and headers), delete data rows from row 5 onwards
        if budget_sheet.max_row > 4:
            budget_sheet.delete_rows(5, budget_sheet.max_row - 4)
            print(f"Cleared data rows from XCC_BUDGET_INTERFACE sheet, keeping header structure")
        
        # Save the changes
        wb.save(output_path)
        wb.close()
    else:
        wb.close()
        raise ValueError("XCC_BUDGET_INTERFACE sheet not found in template")
    
    return str(output_path)

def fill_budget_template_with_data(
    template_path: str, 
    budget_data: List[Dict[str, Any]], 
    output_path: str = None,
    auto_zip: bool = False
) -> str:
    """
    Fill a clean budget template with data and optionally create ZIP.
    
    Args:
        template_path: Path to the clean BudgetImportTemplate.xlsm
        budget_data: List of dictionaries containing budget entry data
        output_path: Path for the filled template (optional)
        auto_zip: Whether to automatically create ZIP file
    
    Returns:
        Path to the filled template file (or ZIP file if auto_zip=True)
    """
    template_file = Path(template_path)
    
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    if not budget_data:
        raise ValueError("No budget data provided")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"BudgetImport_Filled_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)
    
    # Create a copy for filling
    shutil.copy2(template_file, output_path)
    
    # Open the workbook and get the XCC_BUDGET_INTERFACE sheet
    wb = load_workbook(output_path)
    budget_sheet = wb["XCC_BUDGET_INTERFACE"]
    
    # Get the headers from row 4
    headers = []
    for col_num in range(1, budget_sheet.max_column + 1):
        header_cell = budget_sheet.cell(row=4, column=col_num)
        if header_cell.value is not None and str(header_cell.value).strip():
            clean_header = str(header_cell.value).lstrip('*').strip()
            headers.append(clean_header)
        else:
            break
    
    print(f"Found {len(headers)} headers in template")
    print(f"Filling template with {len(budget_data)} budget entries")
    
    # Fill data starting from row 5
    for row_idx, entry in enumerate(budget_data, start=5):
        for col_idx, header in enumerate(headers, start=1):
            # Get value from entry data
            value = entry.get(header, "")
            
            # Set the cell value
            cell = budget_sheet.cell(row=row_idx, column=col_idx)
            
            # Handle different data types appropriately
            if value is not None:
                cell.value = value
            else:
                cell.value = ""  # Empty for None
    
    # Save the filled template
    wb.save(output_path)
    wb.close()
    
    print(f"Template filled with data: {output_path}")
    
    # Optionally create ZIP file
    if auto_zip:
        zip_path = str(output_path).replace('.xlsm', '.zip')
        try:
            zip_result = excel_to_csv_and_zip(str(output_path), zip_path)
            print(f"ZIP file created: {zip_result}")
            return zip_result
        except Exception as e:
            print(f"Warning: Failed to create ZIP file: {e}")
            return str(output_path)
    
    return str(output_path)

def create_budget_from_scratch(
    template_path: str,
    budget_data: List[Dict[str, Any]],
    output_name: str = None,
    auto_zip: bool = True
) -> str:
    """
    Complete workflow: Create clean template, fill with data, and optionally ZIP.
    
    Args:
        template_path: Path to the original BudgetImportTemplate.xlsm
        budget_data: List of dictionaries containing budget entry data
        output_name: Base name for output files (optional)
        auto_zip: Whether to create ZIP file automatically
    
    Returns:
        Path to the final file (Excel or ZIP)
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if output_name is None:
        output_name = f"BudgetImport_{timestamp}"
    
    try:
        # Step 1: Create clean template
        clean_template = create_clean_budget_template(template_path)
        
        # Step 2: Fill with data
        filled_template_path = Path(template_path).parent / f"{output_name}.xlsm"
        result_path = fill_budget_template_with_data(
            clean_template, 
            budget_data, 
            str(filled_template_path),
            auto_zip=auto_zip
        )
        
        # Clean up temporary clean template
        Path(clean_template).unlink(missing_ok=True)
        
        return result_path
    
    except Exception as e:
        print(f"Error in budget creation workflow: {e}")
        raise

def create_and_upload_budget(
    template_path: str,
    budget_data: List[Dict[str, Any]],
    output_name: str = None,
    upload_to_oracle: bool = True,
    group_id: str = None
) -> Dict[str, Any]:
    """
    Complete workflow: Create budget template, fill with data, ZIP, and upload to Oracle Fusion.
    
    Args:
        template_path: Path to the original BudgetImportTemplate.xlsm
        budget_data: List of dictionaries containing budget entry data
        output_name: Base name for output files (optional)
        upload_to_oracle: Whether to upload to Oracle Fusion via FBDI
        group_id: Optional group ID for Oracle upload
    
    Returns:
        Dictionary with workflow results including file paths and upload status
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if output_name is None:
        output_name = f"BudgetImport_{timestamp}"
    
    try:
        # Step 1: Create budget ZIP file
        result_path = create_budget_from_scratch(
            template_path=template_path,
            budget_data=budget_data,
            output_name=output_name,
            auto_zip=True
        )
        
        workflow_result = {
            "success": True,
            "zip_file": result_path,
            "budget_entries": len(budget_data),
            "timestamp": timestamp
        }
        
        # Step 2: Upload to Oracle Fusion if requested
        if upload_to_oracle:
            print(f"Uploading budget to Oracle Fusion...")
            upload_result = upload_budget_from_zip(result_path, group_id)
            
            workflow_result.update({
                "upload_success": upload_result["success"],
                "upload_error": upload_result.get("error"),
                "request_id": upload_result.get("request_id"),
                "group_id": upload_result.get("group_id"),
                "upload_message": upload_result.get("message")
            })
            
            if upload_result["success"]:
                print(f"✅ Budget uploaded successfully!")
                print(f"Request ID: {upload_result.get('request_id', 'N/A')}")
                print(f"Group ID: {upload_result.get('group_id', 'N/A')}")
            else:
                print(f"❌ Upload failed: {upload_result['error']}")
        else:
            workflow_result.update({
                "upload_success": None,
                "upload_message": "Upload skipped (upload_to_oracle=False)"
            })
        
        return workflow_result
        
    except Exception as e:
        return {
            "success": False,
            "error": f"Budget workflow failed: {str(e)}",
            "timestamp": timestamp
        }

# Example usage function
def create_sample_budget_data(transfers, transactions_id) -> List[Dict[str, Any]]:
    """
    Create sample budget entry data for testing.
    
    Args:
        transfers: List of transfer objects with cost_center_code, account_code, project_code attributes
    
    Returns:
        List of sample budget entries
    """
    # Generate unique values that will be the same for all transfers in this function run
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    batch_name = "MIC_HQ_MONTHILY"
    budget_name = f"BUDGET_TRANSFER_{timestamp}"
    sample_data = []
    line_number = 1
    for transfer in transfers:
        if transfer.from_center>0:
            budget_entry = {
                "Source Budget Type": "Hyperion Planning",
                "Source Budget Name": batch_name,
                "Budget Entry Name": budget_name,
                "Line Number": line_number,
                "Amount": -1 * getattr(transfer, 'from_center') if (hasattr(transfer, 'from_center')) else "",
                "Currency Code": "AED",
                "Period Name": "21-Sep",
                "Segment1": str(transfer.cost_center_code),
                "Segment2": str(transfer.account_code),
                "Segment3": str(transfer.project_code),
                "Comment": f"Budget transfer for transaction {transactions_id}",
            }
        if transfer.to_center>0:
            budget_entry = {
                "Source Budget Type": "Hyperion Planning",
                "Source Budget Name": batch_name,
                "Budget Entry Name": budget_name,
                "Line Number": line_number,
                "Amount": getattr(transfer, 'to_center') if (hasattr(transfer, 'to_center')) else "",
                "Currency Code": "AED",
                "Period Name": "21-Sep",
                "Segment1": str(transfer.cost_center_code),
                "Segment2": str(transfer.account_code),
                "Segment3": str(transfer.project_code),
               "Comment": f"Budget transfer for transaction {transactions_id}",
            }
        line_number += 1
        sample_data.append(budget_entry)
    
    return sample_data

if __name__ == "__main__":
    # Example usage
    template_path = "BudgetImportTemplate.xlsm"
    
    # Create sample data (you would replace this with your real transfer data)
    sample_transfers = [
        type('Transfer', (), {
            'cost_center_code': '100001',
            'account_code': '411100',
            'project_code': '001',
            'from_center': 50000,
            'to_center': 0,
            'total_debit': 50000
        })(),
        type('Transfer', (), {
            'cost_center_code': '100002',
            'account_code': '411100',
            'project_code': '002',
            'from_center': 0,
            'to_center': 50000,
            'total_debit': 50000
        })()
    ]
    
    # Create sample data
    sample_data = create_sample_budget_data(sample_transfers, "TXN123")
    
    try:
        print("=== Budget Creation and Upload Workflow ===")
        print(f"Processing {len(sample_data)} budget entries...")
        
        # Option 1: Create budget and upload to Oracle in one step
        result = create_and_upload_budget(
            template_path=template_path,
            budget_data=sample_data,
            output_name="SampleBudget",
            upload_to_oracle=True,  # Set to False if you just want to create the file
            group_id=f"BUDGET_TEST_{time.strftime('%Y%m%d_%H%M%S')}"
        )
        
        print(f"\n=== Workflow Results ===")
        print(f"Success: {result['success']}")
        if result['success']:
            print(f"ZIP File: {result['zip_file']}")
            print(f"Budget Entries: {result['budget_entries']}")
            if result.get('upload_success'):
                print(f"Upload Status: ✅ Success")
                print(f"Request ID: {result.get('request_id', 'N/A')}")
                print(f"Group ID: {result.get('group_id', 'N/A')}")
            elif result.get('upload_success') is False:
                print(f"Upload Status: ❌ Failed - {result.get('upload_error', 'Unknown error')}")
            else:
                print(f"Upload Status: ⏭️ Skipped")
        else:
            print(f"Error: {result.get('error', 'Unknown error')}")
        
        # Option 2: Just create the file without uploading
        # result = create_budget_from_scratch(
        #     template_path=template_path,
        #     budget_data=sample_data,
        #     output_name=r"test_upload_fbdi\SampleBudgetNoUpload",
        #     auto_zip=True
        # )
        # print(f"File created: {result}")
        
    except Exception as e:
        print(f"Error: {e}")