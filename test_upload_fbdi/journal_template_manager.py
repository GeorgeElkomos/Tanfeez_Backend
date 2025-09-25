import pandas as pd
import openpyxl
from openpyxl import load_workbook
import shutil
from pathlib import Path
import time
from typing import Dict, List, Any
from test_upload_fbdi.zip_fbdi import excel_to_csv_and_zip

def create_clean_journal_template(template_path: str, output_path: str = None) -> str:
    """
    Create a clean copy of the JournalImportTemplate with empty GL_INTERFACE data rows.
    
    Args:
        template_path: Path to the original JournalImportTemplate.xlsm
        output_path: Path for the clean template (optional, auto-generates if not provided)
    
    Returns:
        Path to the clean template file
    """
    template_file = Path(template_path)
    
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"JournalImportTemplate_Clean_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)
    
    # Create a copy of the template
    shutil.copy2(template_file, output_path)
    print(f"Created clean template copy: {output_path}")
    
    # Open the copy and clear GL_INTERFACE data
    wb = load_workbook(output_path)
    
    if "GL_INTERFACE" in wb.sheetnames:
        gl_sheet = wb["GL_INTERFACE"]
        
        # Keep rows 1-4 (instructions and headers), delete data rows from row 5 onwards
        if gl_sheet.max_row > 4:
            gl_sheet.delete_rows(5, gl_sheet.max_row - 4)
            print(f"Cleared data rows from GL_INTERFACE sheet, keeping header structure")
        
        # Save the changes
        wb.save(output_path)
        wb.close()
    
    return str(output_path)

def fill_journal_template_with_data(
    template_path: str, 
    journal_data: List[Dict[str, Any]], 
    output_path: str = None,
    auto_zip: bool = False
) -> str:
    """
    Fill a clean journal template with data and optionally create ZIP.
    
    Args:
        template_path: Path to the clean JournalImportTemplate.xlsm
        journal_data: List of dictionaries containing journal entry data
        output_path: Path for the filled template (optional)
        auto_zip: Whether to automatically create ZIP file
    
    Returns:
        Path to the filled template file (or ZIP file if auto_zip=True)
    """
    template_file = Path(template_path)
    
    if not template_file.exists():
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    if not journal_data:
        raise ValueError("No journal data provided")
    
    # Generate output path if not provided
    if output_path is None:
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        output_path = template_file.parent / f"JournalImport_Filled_{timestamp}.xlsm"
    else:
        output_path = Path(output_path)
    
    # Create a copy for filling
    shutil.copy2(template_file, output_path)
    
    # Open the workbook and get the GL_INTERFACE sheet
    wb = load_workbook(output_path)
    gl_sheet = wb["GL_INTERFACE"]
    
    # Get the headers from row 4
    headers = []
    for col_num in range(1, gl_sheet.max_column + 1):
        header_cell = gl_sheet.cell(row=4, column=col_num)
        if header_cell.value is not None and str(header_cell.value).strip():
            clean_header = str(header_cell.value).lstrip('*').strip()
            headers.append(clean_header)
        else:
            break
    
    print(f"Found {len(headers)} headers in template")
    print(f"Filling template with {len(journal_data)} journal entries")
    
    # Fill data starting from row 5
    for row_idx, entry in enumerate(journal_data, start=5):
        for col_idx, header in enumerate(headers, start=1):
            # Get value from entry data
            value = entry.get(header, "")
            
            # Set the cell value
            cell = gl_sheet.cell(row=row_idx, column=col_idx)
            
            # Handle different data types appropriately
            if value is not None:
                cell.value = value
            else:
                cell.value = ""  # Empty for None
            # elif isinstance(value, str) and value.strip():
            #     cell.value = value.strip()
            # else:
            #     cell.value = ""  # Empty for None or empty strings
    
    # Save the filled template
    wb.save(output_path)
    wb.close()
    
    print(f"Template filled with data: {output_path}")
    
    # Optionally create ZIP file
    if auto_zip:
        zip_path = str(output_path).replace('.xlsm', '.zip')
        try:
            zip_result = excel_to_csv_and_zip(str(output_path), zip_path)
            print(f"ZIP file created: {zip_result}")
            return zip_result
        except Exception as e:
            print(f"Warning: Failed to create ZIP file: {e}")
            return str(output_path)
    
    return str(output_path)

def create_journal_from_scratch(
    template_path: str,
    journal_data: List[Dict[str, Any]],
    output_name: str = None,
    auto_zip: bool = True
) -> str:
    """
    Complete workflow: Create clean template, fill with data, and optionally ZIP.
    
    Args:
        template_path: Path to the original JournalImportTemplate.xlsm
        journal_data: List of dictionaries containing journal entry data
        output_name: Base name for output files (optional)
        auto_zip: Whether to create ZIP file automatically
    
    Returns:
        Path to the final file (Excel or ZIP)
    """
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if output_name is None:
        output_name = f"JournalImport_{timestamp}"
    
    try:
        # Step 1: Create clean template
        clean_template = create_clean_journal_template(template_path)
        
        # Step 2: Fill with data
        filled_template_path = Path(template_path).parent / f"{output_name}.xlsm"
        result_path = fill_journal_template_with_data(
            clean_template, 
            journal_data, 
            str(filled_template_path),
            auto_zip=auto_zip
        )
        
        # Clean up temporary clean template
        Path(clean_template).unlink(missing_ok=True)
        
        return result_path
    
    except Exception as e:
        print(f"Error in journal creation workflow: {e}")
        raise

# Example usage function
def create_sample_journal_data(transfers,transaction_id=0,type="submit",group_id=0) -> List[Dict[str, Any]]:
    """
    Create sample journal entry data for testing.
    
    Args:
        transfers: List of transfer objects with cost_center_code, account_code, project_code attributes
    
    Returns:
        List of sample journal entries
    """
    # Generate unique values that will be the same for all transfers in this function run
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    batch_name = f"BATCH_TRANSFER_{timestamp} transaction id={transaction_id}"
    batch_description = f"Balance Transfer Batch created on {time.strftime('%Y-%m-%d %H:%M:%S')}"
    journal_name = f"JOURNAL_TRANSFER_{timestamp} transaction id={transaction_id}"
    journal_description = f"Journal Entry for Balance Transfer - Created {time.strftime('%Y-%m-%d %H:%M:%S')}"

    sample_data = []
    total_debit=0
    for transfer in transfers:
        total_debit += getattr(transfer, 'from_center') if (transfer.from_center is not None) else 0

    for transfer in transfers:
        if transfer.from_center >0:
            journal_entry = {
                "Status Code": "NEW",
                "Ledger ID": "300000205309206",
                "Effective Date of Transaction": "2025-09-17",
                "Journal Source": "Allocations",
                "Journal Category": "Adjustment",
                "Currency Code": "AED",
                "Journal Entry Creation Date": "2025-09-17",
                "Actual Flag": "E",
                "Segment1": transfer.cost_center_code,
                "Segment2": "B040009",
                "Segment3": transfer.account_code,
                "Segment4": "M0000",
                "Segment5": transfer.project_code,
                "Segment6": "00000",
                "Segment7": "000000",
                "Segment8": "000000",
                "Segment9": "000000",
                "Entered Debit Amount": (
                    getattr(transfer, "from_center")
                    if (transfer.from_center is not None) and (type == "submit")
                    else ""
                ),
                "Entered Credit Amount": (
                    getattr(transfer, "from_center")
                    if (transfer.from_center is not None) and (type == "reject")
                    else ""
                ),
                "REFERENCE1 (Batch Name)": batch_name,
                "REFERENCE2 (Batch Description)": batch_description,
                "REFERENCE4 (Journal Entry Name)": journal_name,
                "REFERENCE5 (Journal Entry Description)": journal_description,
                "REFERENCE10 (Journal Entry Line Description)": f"Credit line for account {transfer.account_code}",
                "Encumbrance Type ID": "100000243328511",
                "Interface Group Identifier": group_id,
            }
            sample_data.append(journal_entry)

    journal_entry = {
                "Status Code": "NEW",
                "Ledger ID": "300000205309206",
                "Effective Date of Transaction": "2025-09-17",
                "Journal Source": "Allocations",
                "Journal Category": "Adjustment",
                "Currency Code": "AED",
                "Journal Entry Creation Date": "2025-09-17",
                "Actual Flag": "E",
                "Segment1": 10001,
                "Segment2": "0000000",
                "Segment3": 2205403,
                "Segment4": "M0000",
                "Segment5": "CTRLCE1",
                "Segment6": "00000",
                "Segment7": "000000",
                "Segment8": "000000",
                "Segment9": "000000",
                "Entered Debit Amount": total_debit  if (type=="reject") else "",
                "Entered Credit Amount": total_debit  if (type=="submit") else "",
                "REFERENCE1 (Batch Name)": batch_name,
                "REFERENCE2 (Batch Description)": batch_description,
                "REFERENCE4 (Journal Entry Name)": journal_name,
                "REFERENCE5 (Journal Entry Description)": journal_description,
                "REFERENCE10 (Journal Entry Line Description)": f"Debit line for account {transfer.account_code}",
                "Encumbrance Type ID": "100000243328511",
                "Interface Group Identifier": group_id
            }
    sample_data.append(journal_entry)

    return sample_data

if __name__ == "__main__":
    # Example usage
    template_path = "JournalImportTemplate.xlsm"
    
    # Create sample data
    sample_data = create_sample_journal_data()
    
    try:
        # Create journal from scratch with sample data
        result = create_journal_from_scratch(
            template_path=template_path,
            journal_data=sample_data,
            output_name=r"test_upload_fbdi\SampleJournal",
            auto_zip=True
        )
        
        print(f"\nCompleted! Final file: {result}")
        
    except Exception as e:
        print(f"Error: {e}")
